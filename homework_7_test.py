import os
import pytest
import zipfile
import csv
from zipfile import ZipFile
from openpyxl import load_workbook
from pypdf import PdfReader

current_dir = os.path.dirname(__file__)
dir_path = os.path.join(current_dir, 'tmp')
archive_path = os.path.join(dir_path, "archive_hw7.zip")
files_dir = [i for i in os.listdir(dir_path)]


@pytest.fixture(scope='session', autouse=True)
def archive_manager():
    with zipfile.ZipFile(archive_path, 'w') as zfile:
        for file in files_dir:
            add_file = os.path.join(dir_path, file)
            zfile.write(add_file, os.path.basename(add_file))

    yield
    os.remove(archive_path)


# Запаковать кодом в zip архив несколько разных файлов: pdf, xlsx, csv
def test_archive_list():
    with ZipFile(archive_path) as zfile:
        assert zfile.namelist() == files_dir


def test_xlsx_content():
    with zipfile.ZipFile(archive_path) as zfile:
        with zfile.open('file_example_XLSX_50.xlsx') as xlsx_file:
            workbook = load_workbook(xlsx_file)
            sheet = workbook.active
            cell_35e_content = sheet.cell(row=35, column=5).value
            assert "France" == cell_35e_content


def test_csv_content():
    with zipfile.ZipFile(archive_path) as zfile:
        with zfile.open("csv_file.csv") as csv_file:
            content = csv_file.read().decode('utf-8')
            csv_reader = list(csv.reader(content.splitlines(), delimiter=';'))
            row_4 = csv_reader[3]
            assert row_4[0] == 'Hallo'


def test_pdf_content():
    with zipfile.ZipFile(archive_path) as zfile:
        with zfile.open('Python Testing with Pytest (Brian Okken).pdf') as pdf_file:
            reader = PdfReader(pdf_file)
            text = reader.get_page(1).extract_text()
            assert "This book is licensed to" in text
